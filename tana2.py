import openpyxl as px
import datetime

wb = px.load_workbook("tana_test.xlsx")
#スタートの段階で仕入れを入力か、棚卸しの入力かを選択する

name = input('入力する品目 : ')
qty = float(input('数量を入力 : '))
#name = "しんしん"
#qty = 1.3


num =3
ws = wb["棚卸し表"]
#data = ws['D3'].value#セルの中身の読み込み
max_r = ws.max_row#何行目まであるか取得
#品名を格納するリスト
foods = []
drinks = []
goods = []
now = datetime.datetime.now().strftime("%Y/%m/%d")
name_list = ws["D3":max_r]

for i in range(num,max_r+1):

    if ws["B"+str(i)].value == "食材":
        food_name = ws["D"+str(i)].value
        #print(food_name)
        food = ws["D3":"D"+str(i)]#この範囲で品名が取れる
        #print("{}番目は食材です".format(i))
        last_food_num = i
    elif ws["B"+str(i)].value == "料飲":
        drink_name = ws["D"+str(i)].value
        #print(drink_name)
        drink = ws["D"+str(last_food_num+1):"D"+str(i)]#この範囲で品名が取れる
        #print("{}番目は料飲です".format(i))
        last_drink_num = i
    elif ws["B"+str(i)].value == "消耗品":
        goods_name = ws["D"+str(i)].value
        #print(goods_name)
        good = ws["D"+str(last_drink_num+1):"D"+str(i)]#この範囲で品名が取れる
        #print("{}番目は消耗品です".format(i))
        last_goods_num = i
"""
print("食材は{}列目まで".format(last_food_num))
print("料飲は{}列目まで".format(last_drink_num))
print("消耗品は{}列目まで".format(last_goods_num))
"""
parameter = 0
drink_parameter = 0
good_parameter = 0

#入力された名前が既存のものであるかどうか
for name_i in name_list:
    for name_j in name_i:
        position = name_j.coordinate#セルの位置の特定 "D3"とかが返ってくる
        name_v = ws[position].value
        if name == name_v:
            num_row = position[1:]
            coordi = "I"+str(num_row)
            tana_v = ws[coordi].value
            new_tana_v = tana_v + qty
            ws[coordi].value = new_tana_v
            parameter = 1
            ws["K"+str(num_row)].value = now#更新日
            print("{} の棚卸し数変更前: {} 変更後:{}".format(name_v,tana_v,new_tana_v))

        #foods.append(ws[position].value)

if parameter ==0:
    print('----------------')
    print('{}がなかったので新規追加します'.format(name))
    tag = input('入力する区分 : ')
    price = int(input('単価の入力 : '))
    unit = input("単位を入力 : ")
    if tag == "食材":
        ws.insert_rows(last_food_num+1)
        ws["A"+str(last_food_num+1)].value = "=ROW()-2"
        ws["B"+str(last_food_num+1)].value = tag#区分
        ws["D"+str(last_food_num+1)].value = name#品名
        ws["E"+str(last_food_num+1)].value = price#単価
        ws["F"+str(last_food_num+1)].value = unit#単位
        ws["I"+str(last_food_num+1)].value = qty#量
        ws["K"+str(last_food_num+1)].value = now#更新日
        print('{}行目に挿入しました'.format(last_food_num+1))
    elif tag == "料飲":
        ws.insert_rows(last_drink_num+1)
        ws["A"+str(last_drink_num+1)].value = "=ROW()-2"
        ws["B"+str(last_drink_num+1)].value = tag#区分
        ws["D"+str(last_drink_num+1)].value = name#品名
        ws["E"+str(last_drink_num+1)].value = price#単価
        ws["F"+str(last_drink_num+1)].value = unit#単位
        ws["I"+str(last_drink_num+1)].value = qty#量
        ws["K"+str(last_drink_num+1)].value = now#更新日
        print('{}行目に挿入しました'.format(last_drink_num+1))
    elif tag == "消耗品":
        ws.insert_rows(last_goods_num+1)
        ws["A"+str(last_goods_num+1)].value = "=ROW()-2"
        ws["B"+str(last_goods_num+1)].value = tag#区分
        ws["D"+str(last_goods_num+1)].value = name#品名
        ws["E"+str(last_goods_num+1)].value = price#単価
        ws["F"+str(last_goods_num+1)].value = unit#単位
        ws["I"+str(last_goods_num+1)].value = qty#量
        ws["K"+str(last_goods_num+1)].value = now#更新日
        print('{}行目に挿入しました'.format(last_goods_num+1))


wb.save("tana_test.xlsx")#書き込み
print('保存しました')
